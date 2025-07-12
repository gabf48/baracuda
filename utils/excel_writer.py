from openpyxl import load_workbook

class ExcelWriter:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.sheet = self.wb["Template"]
        self.start_row = 6

    def write_details(self, sku, name, description, index):
        row = self.start_row + index
        self.sheet.cell(row=row, column=1).value = sku
        self.sheet.cell(row=row, column=16).value = name
        self.sheet.cell(row=row, column=17).value = 'Baracuda'
        self.sheet.cell(row=row, column=18).value = description
        self.wb.save(self.file_path)

    def write_image_links(self, links, index):
        """
        Scrie până la 6 linkuri pe coloanele U-Z (21-26),
        în rândul calculat după start_row + index
        """
        row = self.start_row + index
        max_cols = 6
        base_col = 21  # coloana U

        for i in range(min(len(links), max_cols)):
            self.sheet.cell(row=row, column=base_col + i).value = links[i]
        self.wb.save(self.file_path)
